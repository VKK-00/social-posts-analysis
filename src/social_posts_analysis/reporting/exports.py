from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import Workbook

from social_posts_analysis.paths import ProjectPaths


def write_tabular_exports(paths: ProjectPaths, run_id: str, export_tables: dict[str, pl.DataFrame]) -> list[Path]:
    export_root = paths.reports_root / f"report_{run_id}_tables"
    export_root.mkdir(parents=True, exist_ok=True)

    csv_paths: list[Path] = []
    for table_name, frame in export_tables.items():
        csv_path = export_root / f"{table_name}.csv"
        sanitize_export_frame(frame).write_csv(csv_path)
        csv_paths.append(csv_path)

    workbook_path = paths.reports_root / f"report_{run_id}.xlsx"
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    for table_name, frame in export_tables.items():
        worksheet = workbook.create_sheet(title=sheet_name(table_name))
        sanitized = sanitize_export_frame(frame)
        worksheet.append(list(sanitized.columns))
        for row in sanitized.iter_rows():
            worksheet.append([excel_cell_value(value) for value in row])
    workbook.save(workbook_path)
    return [*csv_paths, workbook_path]


def rows_to_frame(rows: list[dict[str, Any]]) -> pl.DataFrame:
    return pl.DataFrame(rows) if rows else pl.DataFrame()


def sanitize_export_frame(frame: pl.DataFrame) -> pl.DataFrame:
    if frame.is_empty():
        return frame
    expressions: list[pl.Expr] = []
    for column_name, dtype in frame.schema.items():
        base_type = dtype.base_type() if hasattr(dtype, "base_type") else None
        if base_type is pl.List:
            expressions.append(
                pl.col(column_name).map_elements(
                    json_list_cell,
                    return_dtype=pl.String,
                ).alias(column_name)
            )
        elif base_type is pl.Struct:
            expressions.append(
                pl.col(column_name).map_elements(
                    json_object_cell,
                    return_dtype=pl.String,
                ).alias(column_name)
            )
        else:
            expressions.append(pl.col(column_name))
    return frame.select(expressions)


def merge_existing_export(path: Path, current: pl.DataFrame, keys: list[str], editable_columns: list[str]) -> pl.DataFrame:
    if not path.exists():
        return current
    existing = pl.read_csv(path)
    editable_columns = [column for column in editable_columns if column in existing.columns and column in current.columns]
    if not editable_columns:
        return current
    merged = current.join(existing.select([*keys, *editable_columns]), on=keys, how="left", suffix="_existing")
    for column in editable_columns:
        if f"{column}_existing" in merged.columns:
            merged = merged.with_columns(
                pl.when(pl.col(f"{column}_existing").is_not_null() & (pl.col(f"{column}_existing") != ""))
                .then(pl.col(f"{column}_existing"))
                .otherwise(pl.col(column))
                .alias(column)
            ).drop(f"{column}_existing")
    return merged


def sheet_name(name: str) -> str:
    trimmed = name[:31]
    return trimmed or "sheet"


def excel_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def json_list_cell(value: Any) -> str:
    if value is None:
        return "[]"
    if isinstance(value, pl.Series):
        return json.dumps(value.to_list(), ensure_ascii=False)
    if isinstance(value, tuple):
        return json.dumps(list(value), ensure_ascii=False)
    return json.dumps(value, ensure_ascii=False)


def json_object_cell(value: Any) -> str:
    if value is None:
        return "{}"
    if isinstance(value, pl.Series):
        return json.dumps(value.to_list(), ensure_ascii=False)
    return json.dumps(value, ensure_ascii=False)

