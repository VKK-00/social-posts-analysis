from __future__ import annotations

import polars as pl
from openpyxl import load_workbook

from facebook_posts_analysis.analysis.service import AnalysisService
from facebook_posts_analysis.normalize import NormalizationService
from facebook_posts_analysis.reporting.service import ReportService, ReviewExportService


def test_report_service_applies_manual_overrides(project_config, project_paths) -> None:
    NormalizationService(project_config, project_paths).run(run_id="20260402T120000Z")
    AnalysisService(project_config, project_paths).run(run_id="20260402T120000Z")
    ReviewExportService(project_config, project_paths).run(run_id="20260402T120000Z")

    narrative_overrides = pl.read_csv(project_paths.review_root / "narrative_overrides.csv")
    comment_cluster_ids = narrative_overrides.filter(pl.col("item_type") == "comment")["cluster_id"].to_list()
    if comment_cluster_ids:
        target_cluster_id = comment_cluster_ids[0]
        narrative_overrides = narrative_overrides.with_columns(
            pl.when((pl.col("item_type") == "comment") & (pl.col("cluster_id") == target_cluster_id))
            .then(pl.lit("Renamed narrative"))
            .otherwise(pl.col("new_label"))
            .alias("new_label")
        )
        narrative_overrides.write_csv(project_paths.review_root / "narrative_overrides.csv")

    stance_overrides = pl.read_csv(project_paths.review_root / "stance_overrides.csv")
    side_id = project_config.sides[0].side_id
    stance_overrides = stance_overrides.with_columns(
        pl.when((pl.col("item_id") == "comment_1") & (pl.col("side_id") == side_id))
        .then(pl.lit("oppose"))
        .otherwise(pl.col("override_label"))
        .alias("override_label")
    )
    stance_overrides.write_csv(project_paths.review_root / "stance_overrides.csv")

    context = ReportService(project_config, project_paths)._build_context("20260402T120000Z")

    assert context["global_support"]
    side_row = next(row for row in context["global_support"] if row["side_id"] == side_id)
    assert side_row["oppose_count"] >= 1
    if context["comment_clusters"]:
        assert any(cluster["label"] == "Renamed narrative" for cluster in context["comment_clusters"])


def test_report_service_writes_tabular_exports(project_config, project_paths) -> None:
    NormalizationService(project_config, project_paths).run(run_id="20260402T120000Z")
    AnalysisService(project_config, project_paths).run(run_id="20260402T120000Z")

    outputs = ReportService(project_config, project_paths).run(run_id="20260402T120000Z")

    workbook_path = project_paths.reports_root / "report_20260402T120000Z.xlsx"
    export_root = project_paths.reports_root / "report_20260402T120000Z_tables"
    assert workbook_path in outputs
    assert (export_root / "per_side_support.csv").exists()
    assert (export_root / "per_post_overview.csv").exists()

    workbook = load_workbook(workbook_path)
    assert "per_side_support" in workbook.sheetnames
    assert "per_post_overview" in workbook.sheetnames
